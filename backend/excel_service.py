from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime
from typing import List


class ExcelExportService:
    """Excel raporu oluşturma servisi"""
    
    def generate_orders_report(self, orders: List[dict], title: str = "Sipariş Raporu") -> bytes:
        """Sipariş raporunu Excel olarak oluştur"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Siparişler"
        
        # Başlık
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:G1')
        
        # Sütun başlıkları
        headers = ['Sipariş No', 'Tarih', 'Tip', 'Müşteri', 'Tutar', 'Durum', 'Kurye']
        header_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Veri satırları
        for row_idx, order in enumerate(orders, start=4):
            ws.cell(row=row_idx, column=1, value=order.get('order_number', 'N/A'))
            
            created_at = order.get('created_at')
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at)
            ws.cell(row=row_idx, column=2, value=created_at.strftime('%d.%m.%Y %H:%M') if created_at else 'N/A')
            
            order_type_map = {'dine-in': 'İçeride', 'takeaway': 'Paket', 'delivery': 'Gel-Al'}
            ws.cell(row=row_idx, column=3, value=order_type_map.get(order.get('order_type'), 'N/A'))
            
            ws.cell(row=row_idx, column=4, value=order.get('customer_name', 'N/A'))
            ws.cell(row=row_idx, column=5, value=f"{order.get('total_amount', 0):.2f} ₺")
            
            status_map = {
                'pending': 'Bekliyor',
                'preparing': 'Hazırlanıyor',
                'ready': 'Hazır',
                'delivered': 'Teslim Edildi',
                'cancelled': 'İptal'
            }
            ws.cell(row=row_idx, column=6, value=status_map.get(order.get('status'), 'N/A'))
            ws.cell(row=row_idx, column=7, value=order.get('courier_name', '-'))
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        # BytesIO'ya kaydet
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()